# -*- coding: utf-8 -*-
"""
PlaceExplorer - version GitHub Actions
- Lit la localisation et le mois depuis les variables d'environnement (inputs du workflow)
- Google Places API (New) : 1 seul appel searchText par catégorie (32 catégories,
  top 20 lieux avec note, avis, prix, horaires, site web, résumé éditorial)
- Enrichissement Wikipédia gratuit (article + consultations mensuelles) pour
  booster les incontournables dans l'itinéraire
- Génère l'Excel multi-feuilles et envoie un email HTML stylé depuis
  romtaug@gmail.com avec l'Excel (+ images si présentes) en pièce jointe

Secrets requis (GitHub > Settings > Secrets and variables > Actions) :
- GOOGLE_API_KEY      : clé API Google (avec "Places API (New)" ACTIVÉE dans GCP)
- GMAIL_APP_PASSWORD  : mot de passe d'application Gmail de romtaug@gmail.com
"""

import os
import re
import sys
import time
import shutil
import smtplib
import unicodedata
import datetime as dt
from urllib.parse import quote

import requests
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email import encoders
from openpyxl import load_workbook

from maps_route import (build_day_itinerary, build_multiday_plan, itinerary_email_block,
                        itinerary_plain_block, describe, download_static_map)

# ----------------------------------------------------------------------------
# Configuration (depuis l'environnement - injecté par le workflow)
# ----------------------------------------------------------------------------
API_KEY = os.environ.get("GOOGLE_API_KEY", "")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
SENDER_EMAIL = "romtaug@gmail.com"
RECEIVER_EMAILS = [e.strip() for e in os.environ.get("RECEIVER_EMAILS", "romtaug@gmail.com").split(",") if e.strip()]

LOCATION = os.environ.get("LOCATION", "").strip()
VACATION_MONTH = os.environ.get("VACATION_MONTH", "").strip()
try:
    NB_DAYS = int(os.environ.get("NB_DAYS", "0") or "0")
except ValueError:
    NB_DAYS = 0

if not API_KEY:
    print("❌ GOOGLE_API_KEY manquant (secret GitHub).")
    sys.exit(1)
if not GMAIL_APP_PASSWORD:
    print("❌ GMAIL_APP_PASSWORD manquant (secret GitHub).")
    sys.exit(1)
if not LOCATION or not VACATION_MONTH:
    print("❌ LOCATION ou VACATION_MONTH manquant (inputs du workflow).")
    sys.exit(1)


def is_valid_email(email):
    return re.match(r"[^@]+@[^@]+\.[^@]+", email)


for email in RECEIVER_EMAILS:
    if not is_valid_email(email):
        print(f"❌ Adresse email invalide : {email}")
        sys.exit(1)
print("✅ Toutes les adresses email sont valides.")


# ----------------------------------------------------------------------------
# Normalisation
# ----------------------------------------------------------------------------
def normalize_location(location):
    """Force le format 'Ville, Pays' ou 'Pays'. Échoue proprement en CI si invalide."""
    location = " ".join(location.strip().split())
    if "," in location:
        parts = location.split(",")
        if len(parts) == 2:
            city, country = parts[0].strip(), parts[1].strip()
            if city and country:
                return f"{city.title()}, {country.title()}"
    elif location.replace(" ", "").isalpha():
        return location.title()
    print("❌ Format incorrect. Utilisez 'Ville, Pays' ou 'Pays' (ex : Paris, France).")
    sys.exit(1)


def normalize_filename(filename):
    return ''.join(
        c for c in unicodedata.normalize('NFD', filename)
        if unicodedata.category(c) != 'Mn'
    )


def extract_city_from_address(full_address):
    address_parts = full_address.split(", ")
    if "Unnamed Road" in full_address:
        city = address_parts[-2] if len(address_parts) > 2 else address_parts[-1]
    elif len(address_parts) > 2:
        city = address_parts[-2]
    elif len(address_parts) == 2:
        city = address_parts[0]
    else:
        city = full_address.split(" ")[0]
    return city


# ----------------------------------------------------------------------------
# Google Places API (New) - https://places.googleapis.com/v1/places:searchText
# UN SEUL appel par catégorie : le field mask ramène TOUTES les infos (note,
# avis, prix, horaires, site web, résumé éditorial, types) pour les 20 lieux
# d'un coup. Fini les ~600 appels Place Details du legacy (~632 -> ~32 appels
# par exécution, largement dans le palier gratuit mensuel du SKU).
# Prérequis console GCP : activer "Places API (New)" et l'autoriser sur la clé.
# ----------------------------------------------------------------------------
PLACES_FIELD_MASK = ",".join(f"places.{f}" for f in (
    "id", "displayName", "formattedAddress", "addressComponents", "location",
    "types", "primaryType", "primaryTypeDisplayName", "businessStatus",
    "rating", "userRatingCount", "priceLevel", "internationalPhoneNumber",
    "websiteUri", "regularOpeningHours", "editorialSummary",
))

_PRICE_LEVELS = {
    'PRICE_LEVEL_FREE': 'Free', 'PRICE_LEVEL_INEXPENSIVE': '+',
    'PRICE_LEVEL_MODERATE': '++', 'PRICE_LEVEL_EXPENSIVE': '+++',
    'PRICE_LEVEL_VERY_EXPENSIVE': '++++',
}

_EN_FR_DAYS = [('monday', 'lundi'), ('tuesday', 'mardi'), ('wednesday', 'mercredi'),
               ('thursday', 'jeudi'), ('friday', 'vendredi'), ('saturday', 'samedi'),
               ('sunday', 'dimanche')]


def _closed_days_fr(opening):
    """Jours de fermeture hebdomadaires, en français, depuis regularOpeningHours.
    Parse les weekdayDescriptions ('lundi: Fermé' en fr, 'Monday: Closed' en en) :
    plus robuste que d'interpréter la convention des periods, et bilingue pour
    survivre à un changement de languageCode. Vide si ouvert 7j/7 ou inconnu."""
    descs = (opening or {}).get('weekdayDescriptions') or []
    closed = set()
    for d in descs:
        low = _norm_txt(d)
        if 'closed' in low or 'ferme' in low:
            closed.add(low.split(':', 1)[0].strip())
    return ', '.join(fr for en, fr in _EN_FR_DAYS if en in closed or fr in closed)


def _extract_country(components):
    """Renvoie (pays_long, code_pays) depuis addressComponents (New API :
    longText/shortText), ex. ('France', 'FR')."""
    for comp in (components or []):
        if 'country' in (comp.get('types') or []):
            return comp.get('longText'), comp.get('shortText')
    return None, None


def _compose_description(pl):
    """Description de base, enrichie au maximum SANS appel supplémentaire :
    drapeau de fermeture temporaire + type de lieu localisé (ex. 'Site
    historique') + résumé éditorial Google. L'introduction Wikipédia sera
    fusionnée ensuite (finalize_descriptions) pour les lieux reconnus."""
    type_label = (pl.get('primaryTypeDisplayName') or {}).get('text', '')
    editorial = " ".join(((pl.get('editorialSummary') or {}).get('text') or '').split())
    parts = []
    if pl.get('businessStatus') == 'CLOSED_TEMPORARILY':
        parts.append("⚠️ Fermé temporairement")
    if type_label and _norm_txt(type_label) not in _norm_txt(editorial):
        parts.append(type_label)
    if editorial:
        parts.append(editorial)
    return " · ".join(parts)


def _parse_new_place(pl):
    """Transforme un objet 'place' de l'API New en dict interne de l'outil.
    Renvoie None pour les lieux fermés définitivement (ni Excel ni parcours)."""
    place_id = pl.get('id')
    if not place_id:
        return None
    if pl.get('businessStatus') == 'CLOSED_PERMANENTLY':
        name = (pl.get('displayName') or {}).get('text', place_id)
        print(f"   ⏭️ Fermé définitivement, ignoré : {name}")
        return None
    loc = pl.get('location') or {}
    full_address = pl.get('formattedAddress', 'Not specified')
    country_long, country_short = _extract_country(pl.get('addressComponents'))
    website = pl.get('websiteUri')
    return {
        'City': extract_city_from_address(full_address),
        'Address': full_address,
        'Name': (pl.get('displayName') or {}).get('text', 'Not specified'),
        'Total Reviews': pl.get('userRatingCount', 0),
        'Rating (on 5)': pl.get('rating', 'Not rated'),
        'Price Level': _PRICE_LEVELS.get(pl.get('priceLevel'), 'Not specified'),
        'Maps': f'=HYPERLINK("https://www.google.com/maps/place/?q=place_id:{place_id}", "📍 Google Maps")',
        'Website': f'=HYPERLINK("{website}", "🌐 Site web")' if website else '',
        'Phone': pl.get('internationalPhoneNumber', 'Not available'),
        # Jours de fermeture hebdo -> colonne Excel + avertissement itinéraire.
        'Fermeture': _closed_days_fr(pl.get('regularOpeningHours')),
        # Description (colonne FINALE de l'Excel), composée par
        # _compose_description ; l'introduction Wikipédia y est fusionnée
        # ensuite (finalize_descriptions), AVANT l'écriture de l'Excel.
        'Description': _compose_description(pl),
        'PlaceID': place_id,
        'Lat': loc.get('latitude'),
        'Lng': loc.get('longitude'),
        'Country': country_long,
        'CountryCode': country_short,
        'Types': pl.get('types') or [],
        'PrimaryType': pl.get('primaryType', ''),
        'BusinessStatus': pl.get('businessStatus', ''),
    }


def geocode_location(location, api_key):
    """Coordonnées (lat, lng) de la ville saisie par l'utilisateur, via Places
    API (New) searchText (l'API Geocoding classique n'est pas activable sur les
    projets GCP récents). Sert d'ancre FIABLE pour l'itinéraire : on se centre
    sur la ville DEMANDÉE, pas sur le lieu le plus visité parmi les résultats
    (qui peut être dans une autre ville si Google élargit la recherche pour une
    catégorie absente localement, ex. Šiauliai -> Klaipėda). None si échec."""
    try:
        r = requests.post(
            "https://places.googleapis.com/v1/places:searchText",
            json={"textQuery": location, "languageCode": "fr", "pageSize": 1},
            headers={"Content-Type": "application/json",
                     "X-Goog-Api-Key": api_key,
                     "X-Goog-FieldMask": "places.location"},
            timeout=30)
        if r.status_code == 200:
            places = (r.json() or {}).get("places", [])
            if places and places[0].get("location"):
                loc = places[0]["location"]
                return (loc["latitude"], loc["longitude"])
            print(f"⚠️ Ancre '{location}' : aucun résultat searchText")
        else:
            print(f"⚠️ Ancre '{location}' : erreur {r.status_code}")
    except Exception as e:
        print(f"⚠️ Ancre erreur : {e}")
    return None


def search_places(api_key, location, category):
    url = "https://places.googleapis.com/v1/places:searchText"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": api_key,
        # Le field mask = la liste des champs facturés ET renvoyés pour CHAQUE
        # lieu du résultat. Un seul appel remplace 1 search + 20 details legacy.
        "X-Goog-FieldMask": PLACES_FIELD_MASK,
    }
    body = {"textQuery": f"{category} in {location}",
            "languageCode": "fr", "pageSize": 20}   # noms, adresses, résumés en FR
    try:
        response = requests.post(url, json=body, headers=headers, timeout=30)
    except Exception as e:
        print(f"⚠️ Réseau KO pour {category} : {e}")
        return []
    if response.status_code != 200:
        detail = ""
        try:
            detail = (response.json().get('error') or {}).get('message', '')
        except Exception:
            pass
        print(f"⚠️ Erreur {response.status_code} pour {category} : {detail}")
        if response.status_code == 403:
            print("   💡 Vérifie que 'Places API (New)' est ACTIVÉE dans la console "
                  "GCP et autorisée sur ta clé (c'est une API distincte du legacy).")
        return []
    places = response.json().get('places', [])
    return [p for p in (_parse_new_place(pl) for pl in places) if p]


# ----------------------------------------------------------------------------
# Enrichissement Wikipédia (GRATUIT, sans clé) - signal de popularité externe
# ----------------------------------------------------------------------------
# Idée (validée par la littérature : Oberholzer et al. 2023, les avis Google
# combinés aux pageviews Wikipédia expliquent le mieux la fréquentation réelle) :
# un lieu qui a un article Wikipédia consulté est un incontournable ; l'article
# est trouvé par GÉOLOCALISATION (geosearch autour des coordonnées du lieu) puis
# validé par similarité de nom -> pas d'homonymes. On enrichit uniquement le
# top des visites (pas les restos/bars), et le score de l'itinéraire booste
# les lieux reconnus. Deux APIs Wikimedia, aucune clé, juste un User-Agent poli.
WIKI_UA = {"User-Agent": "PlaceExplorer/2.0 (https://github.com/romtaug/PlaceExplorer-Claude; romtaug@gmail.com)"}
WIKI_LANGS = ("fr", "en")          # fr d'abord (destinations francophones), puis en


def _wiki_geosearch(lat, lng, lang, radius=400, limit=5):
    """Titres d'articles Wikipédia géolocalisés à moins de `radius` m du lieu."""
    r = requests.get(f"https://{lang}.wikipedia.org/w/api.php", params={
        "action": "query", "list": "geosearch", "gscoord": f"{lat}|{lng}",
        "gsradius": radius, "gslimit": limit, "format": "json"},
        headers=WIKI_UA, timeout=15)
    if r.status_code != 200:
        return []
    return [g.get("title", "") for g in
            (r.json().get("query") or {}).get("geosearch", [])]


def _wiki_title_score(name, title):
    """Similarité nom du lieu <-> titre d'article : part des mots significatifs
    en commun (accents/casse ignorés, mots-outils retirés). 1.0 = match parfait."""
    stop = {"de", "du", "des", "la", "le", "les", "l", "d", "et",
            "of", "the", "a", "an", "and", "in"}
    a = set(_norm_txt(name).replace("'", " ").replace("-", " ").split()) - stop
    b = set(_norm_txt(title).replace("'", " ").replace("-", " ").split()) - stop
    if not a or not b:
        return 0.0
    return len(a & b) / min(len(a), len(b))


def _wiki_extract(lang, title, max_chars=450):
    """Introduction de l'article Wikipédia (2-3 phrases, texte brut).
    C'est le gros de l'enrichissement de la colonne Description : une vraie
    présentation rédigée du lieu, en français quand l'article fr existe."""
    r = requests.get(f"https://{lang}.wikipedia.org/w/api.php", params={
        "action": "query", "prop": "extracts", "exintro": 1, "explaintext": 1,
        "exsentences": 3, "redirects": 1, "titles": title, "format": "json"},
        headers=WIKI_UA, timeout=15)
    if r.status_code != 200:
        return ""
    pages = ((r.json().get("query") or {}).get("pages") or {})
    text = " ".join(next(iter(pages.values()), {}).get("extract", "").split())
    if len(text) > max_chars:                    # coupe à la fin d'une phrase
        cut = text[:max_chars].rsplit(". ", 1)[0]
        text = (cut + "." if cut else text[:max_chars]).strip()
    return text


def _wiki_monthly_views(lang, title):
    """Moyenne mensuelle de consultations de l'article sur ~12 derniers mois
    complets (API REST Wikimedia pageviews)."""
    end = dt.date.today().replace(day=1) - dt.timedelta(days=1)   # fin du mois dernier
    start = (end - dt.timedelta(days=330)).replace(day=1)          # ~12 mois plus tôt
    t = quote(title.replace(" ", "_"), safe="")
    url = (f"https://wikimedia.org/api/rest_v1/metrics/pageviews/per-article/"
           f"{lang}.wikipedia/all-access/all-agents/{t}/monthly/"
           f"{start:%Y%m%d}00/{end:%Y%m%d}00")
    r = requests.get(url, headers=WIKI_UA, timeout=15)
    if r.status_code != 200:
        return 0
    items = (r.json() or {}).get("items", [])
    if not items:
        return 0
    return round(sum(i.get("views", 0) for i in items) / len(items))


def wiki_enrich(place):
    """Tente de relier UN lieu à son article Wikipédia (fr puis en). Si trouvé,
    pose WikiTitle / WikiLang / WikiViews / WikiUrl sur le dict (partagé avec
    l'itinéraire, qui s'en sert pour booster le classement et lier l'article).
    Renvoie True si un article a été reconnu."""
    lat, lng = place.get('Lat'), place.get('Lng')
    if place.get('WikiTitle'):
        # Titre déjà connu (ex. tag wikipedia d'OpenStreetMap) : il ne manque
        # que la popularité de l'article et son introduction.
        if place.get('WikiViews') is None:
            try:
                place['WikiViews'] = _wiki_monthly_views(
                    place.get('WikiLang', 'fr'), place['WikiTitle'])
            except Exception:
                place['WikiViews'] = 0
        if place.get('WikiExtract') is None:
            try:
                place['WikiExtract'] = _wiki_extract(
                    place.get('WikiLang', 'fr'), place['WikiTitle'])
            except Exception:
                place['WikiExtract'] = ''
        return True
    if lat is None or lng is None:
        return False
    for lang in WIKI_LANGS:
        try:
            titles = _wiki_geosearch(lat, lng, lang)
        except Exception:
            titles = []
        best, best_score = None, 0.0
        for t in titles:
            sc = _wiki_title_score(place.get('Name', ''), t)
            if sc > best_score:
                best, best_score = t, sc
        if best and best_score >= 0.5:      # seuil prudent : pas d'homonyme forcé
            try:
                views = _wiki_monthly_views(lang, best)
            except Exception:
                views = 0
            place['WikiTitle'] = best
            place['WikiLang'] = lang
            place['WikiViews'] = views
            place['WikiUrl'] = (f"https://{lang}.wikipedia.org/wiki/"
                                + quote(best.replace(" ", "_")))
            try:
                place['WikiExtract'] = _wiki_extract(lang, best)
            except Exception:
                place['WikiExtract'] = ''
            return True
        time.sleep(0.05)
    return False


def enrich_with_wikipedia(places_by_category, top_n=40):
    """Enrichit le top `top_n` des VISITES potentielles (les dicts étant
    partagés, l'itinéraire voit directement les champs Wiki). Restos et bars
    ne sont pas concernés : le boost ne s'applique qu'aux lieux à visiter."""
    from maps_route import _classify, _dedup, _top
    visits, _, _ = _classify(places_by_category)
    pool = _top(_dedup(visits), top_n)
    hits = 0
    for p in pool:
        if wiki_enrich(p):
            hits += 1
        time.sleep(0.05)                    # politesse API Wikimedia
    print(f"🔎 Wikipédia : {hits}/{len(pool)} lieux reliés à un article "
          f"(boost popularité + lien dans le parcours).")


# ----------------------------------------------------------------------------
# Points de vue OpenStreetMap (GRATUIT, sans clé) - les belvédères que Google
# type mal ou ignore. Une seule requête Overpass autour de la ville, injectée
# UNIQUEMENT dans le parcours (l'Excel reste 100% Google). Fail-safe : toute
# erreur est avalée, OSM_VIEWPOINTS=0 désactive.
# ----------------------------------------------------------------------------
def finalize_descriptions(places_by_category, max_chars=600):
    """Fusionne l'introduction Wikipédia dans la colonne Description, AVANT
    l'écriture de l'Excel. Règles : si l'éditorial Google est redondant avec
    l'extrait (préfixe commun), l'extrait (plus riche) le remplace ; sinon les
    deux sont enchaînés. Coupé proprement à la fin d'une phrase."""
    merged = 0
    for places in places_by_category.values():
        for p in places:
            ext = " ".join((p.get('WikiExtract') or '').split())
            if not ext:
                continue
            desc = (p.get('Description') or '').strip()
            # Redondance : on compare la DERNIÈRE partie de la description
            # (l'éditorial, après 'type ·') à l'extrait, ponctuation ignorée ;
            # si l'extrait la contient déjà, il la remplace (préfixes gardés).
            parts = [x.strip() for x in desc.split('·')] if desc else []
            tail_n = _norm_txt(parts[-1]).rstrip('. ').strip() if parts else ''
            if not desc:
                desc = ext
            elif tail_n and tail_n[:60] in _norm_txt(ext):
                prefix = ' · '.join(parts[:-1])
                desc = (prefix + ' · ' + ext) if prefix else ext
            else:
                desc = desc.rstrip('.') + '. ' + ext
            if len(desc) > max_chars:
                cut = desc[:max_chars].rsplit('. ', 1)[0]
                desc = (cut + '.') if cut else desc[:max_chars]
            p['Description'] = desc
            merged += 1
    if merged:
        print(f"📝 Descriptions enrichies par Wikipédia : {merged} lieu(x).")


def fetch_osm_viewpoints(anchor, radius_km=10, limit=25):
    lat, lng = anchor
    query = (f'[out:json][timeout:15];'
             f'node(around:{int(radius_km * 1000)},{lat},{lng})'
             f'["tourism"="viewpoint"]["name"];'
             f'out body {limit};')
    r = requests.post("https://overpass-api.de/api/interpreter",
                      data={"data": query}, headers=WIKI_UA, timeout=25)
    if r.status_code != 200:
        return []
    out = []
    for el in (r.json() or {}).get("elements", []):
        tags = el.get("tags") or {}
        name = tags.get("name")
        if not name or el.get("lat") is None:
            continue
        place = {
            'City': '', 'Address': tags.get('addr:full', ''),
            'Name': name, 'Total Reviews': 0, 'Rating (on 5)': 'Not rated',
            'Price Level': 'Free',
            'Maps': (f'=HYPERLINK("https://www.google.com/maps/search/?api=1'
                     f'&query={el["lat"]},{el["lon"]}", "📍 Google Maps")'),
            'Website': '', 'Phone': 'Not available',
            'Description': 'Point de vue (OpenStreetMap).', 'Fermeture': '',
            'PlaceID': f'osm_{el.get("id")}',
            'Lat': el.get('lat'), 'Lng': el.get('lon'),
            'Country': None, 'CountryCode': None,
            'Types': ['viewpoint'], 'PrimaryType': 'viewpoint',
            'BusinessStatus': 'OPERATIONAL',
        }
        # Tag wikipedia OSM ('fr:Notre-Dame de la Serra') -> article direct,
        # wiki_enrich complètera juste les vues mensuelles.
        wp = tags.get('wikipedia', '')
        if ':' in wp:
            lang, title = wp.split(':', 1)
            if len(lang) == 2 and title:
                place['WikiLang'], place['WikiTitle'] = lang, title
                place['WikiUrl'] = (f"https://{lang}.wikipedia.org/wiki/"
                                    + quote(title.replace(' ', '_')))
        out.append(place)
    return out


def inject_osm_viewpoints(places_by_category, anchor):
    """Ajoute les belvédères OSM comme pseudo-catégorie du PARCOURS, en
    écartant ceux qui doublonnent un lieu Google (à moins de 150 m avec un
    nom similaire). Les dicts n'entrent pas dans l'Excel (créé avant)."""
    try:
        vps = fetch_osm_viewpoints(anchor)
    except Exception as e:
        print(f"⚠️ Points de vue OSM ignorés : {e}")
        return
    if not vps:
        return
    from maps_route import hav
    existing = [p for pls in places_by_category.values() for p in pls
                if p.get('Lat') is not None]
    kept = []
    for vp in vps:
        dup = any(hav(vp['Lat'], vp['Lng'], e['Lat'], e['Lng']) <= 150
                  and _wiki_title_score(vp['Name'], e.get('Name', '')) >= 0.5
                  for e in existing)
        if not dup:
            vp['_category'] = '🔭 Point de vue'
            kept.append(vp)
    if kept:
        # Les belvédères tagués wikipedia récupèrent vues + extrait
        # (wiki_enrich, branche 'titre déjà connu') pour profiter du boost.
        for vp in kept:
            if vp.get('WikiTitle'):
                wiki_enrich(vp)
        places_by_category['viewpoints'] = kept
        print(f"🔭 OpenStreetMap : {len(kept)} point(s) de vue ajouté(s) au "
              f"parcours ({len(vps) - len(kept)} doublon(s) écarté(s)).")


# ----------------------------------------------------------------------------
# Excel
# ----------------------------------------------------------------------------
def adjust_column_width(file_path):
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(file_path)



def style_workbook(file_path):
    """
    Stylise tout le classeur : en-têtes navy, lignes alternées, filtres,
    volet figé, liens Maps en bleu, formats de nombres, largeurs ajustées.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule

    NAVY = "1F3B63"
    BAND = "F3F6FA"
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    band_fill = PatternFill(start_color=BAND, end_color=BAND, fill_type="solid")
    body_font = Font(name="Calibri", size=10.5, color="3C4043")
    link_font = Font(name="Calibri", size=10.5, color="1A73E8", underline="single", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin_border = Border(bottom=Side(style="thin", color="E0E5EA"))

    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1:
            continue

        headers = [c.value for c in ws[1]]
        col_idx = {h: i + 1 for i, h in enumerate(headers)}

        # En-tête
        ws.row_dimensions[1].height = 24
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Corps
        for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.font = body_font
                cell.border = thin_border
                cell.alignment = left
                if r % 2 == 0:
                    cell.fill = band_fill
            # Colonnes spécifiques
            if 'Total Reviews' in col_idx:
                c = ws.cell(row=r, column=col_idx['Total Reviews'])
                c.number_format = '#,##0'
                c.alignment = center
            for name in ('Rating (on 5)', 'Price Level', 'City'):
                if name in col_idx:
                    ws.cell(row=r, column=col_idx[name]).alignment = center
            for name in ('Maps', 'Website', 'Wikipédia'):
                if name in col_idx:
                    c = ws.cell(row=r, column=col_idx[name])
                    c.font = link_font
                    c.alignment = center

        # Volet figé + filtres
        ws.freeze_panes = "A2"
        if ws.max_row >= 2:
            ws.auto_filter.ref = ws.dimensions

        # Heatmap (dégradé continu) sur le nombre d'avis = popularité.
        # Plus c'est vert, plus le lieu est populaire ; la valeur max ressort.
        if 'Total Reviews' in col_idx and ws.max_row >= 2:
            letter = get_column_letter(col_idx['Total Reviews'])
            rng = f"{letter}2:{letter}{ws.max_row}"
            ws.conditional_formatting.add(rng, ColorScaleRule(
                start_type='min', start_color='FFFFFF',
                mid_type='percentile', mid_value=50, mid_color='B7E1A1',
                end_type='max', end_color='4CAF50'))

        # Largeurs : basées sur le contenu, libellé fixe pour Maps, bornées
        for i, col in enumerate(ws.columns, start=1):
            letter = get_column_letter(i)
            header = headers[i - 1] if i - 1 < len(headers) else ""
            if header == 'Maps':
                ws.column_dimensions[letter].width = 18
                continue
            if header == 'Signification':          # colonne longue de la Légende
                ws.column_dimensions[letter].width = 90
                for cell in col:
                    cell.alignment = Alignment(horizontal="left", vertical="center",
                                               wrap_text=True)
                continue
            if header == 'Description':            # colonne longue finale
                ws.column_dimensions[letter].width = 70
                for cell in col:
                    if cell.row > 1:
                        cell.alignment = Alignment(horizontal="left",
                                                   vertical="top", wrap_text=True)
                continue
            max_length = len(str(header or ""))
            for cell in col:
                if cell.row == 1 or cell.value is None:
                    continue
                max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[letter].width = max(12, min(max_length + 3, 55))

    wb.save(file_path)
    print(f"✅ Classeur stylisé : {file_path}")


def _norm_txt(s):
    """minuscule + sans accents, pour comparer des noms de pays sereinement."""
    s = unicodedata.normalize('NFD', str(s or ''))
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    return s.strip().lower()


def filter_by_country(places_by_category, location):
    """Garde uniquement les lieux situés dans le pays demandé, en s'appuyant sur
    le composant 'country' renvoyé par l'API (et non sur la fin de l'adresse).
    - Tolérant : un lieu sans info pays est conservé.
    - Garde-fou : si AUCUN lieu ne matche (ex. pays mal orthographié ou tapé
      dans une autre langue), on n'applique PAS le filtre plutôt que de vider
      le guide. Les résultats restent pertinents car la requête Places est déjà
      géocodée sur le bon pays.
    """
    requested = _norm_txt(location.split(",")[-1])
    if not requested:
        return places_by_category

    def has_country(p):
        return bool(_norm_txt(p.get('Country')) or _norm_txt(p.get('CountryCode')))

    def matches(p):
        cl, cc = _norm_txt(p.get('Country')), _norm_txt(p.get('CountryCode'))
        return (requested == cl or requested == cc
                or requested in cl or cl in requested)

    all_places = [p for v in places_by_category.values() for p in v]
    total = len(all_places)
    with_country = [p for p in all_places if has_country(p)]
    matched = [p for p in with_country if matches(p)]

    # Garde-fou anti-vidage : si des lieux ont une info pays mais qu'AUCUN ne
    # matche (pays tapé dans une autre langue, ex. 'Japon' vs 'Japan', ou faute
    # de frappe), on n'applique PAS le filtre. La requête Places étant déjà
    # géocodée sur le bon pays, les résultats restent pertinents.
    label = location.split(",")[-1].strip()
    if with_country and not matched:
        print(f"⚠️ Filtre pays '{label}' : 0 correspondance sur les lieux "
              f"géolocalisés -> filtre ignoré (vérifie l'orthographe / la langue).")
        return places_by_category

    # On garde : les lieux qui matchent + ceux sans info pays (on ne les jette pas).
    def ok(p):
        return matches(p) if has_country(p) else True

    kept = {c: [p for p in places if ok(p)] for c, places in places_by_category.items()}
    kept = {c: v for c, v in kept.items() if v}
    kept_total = sum(len(v) for v in kept.values())
    print(f"✅ Filtre pays '{label}' : {kept_total}/{total} lieux conservés.")
    return kept


CATEGORIES = {
    # La requête reine : les incontournables au sens de Google. Alimente une
    # feuille dédiée en tête d'Excel ET enrichit le pool de l'itinéraire avec
    # exactement ce qu'on cherche (au lieu de le reconstituer par ricochet).
    'attractions': '🎯 Incontournables',
    # Découvertes
    'historical_sites': '🏰 Sites historiques',
    'museums': '🖼️ Musées',
    'churches': '⛪ Églises',
    'cultural_centers': '🎭 Centres culturels',
    'hiking_trails': '🥾 Sentiers de randonnée',
    # Restauration
    'restaurants': '🍴 Restaurants',
    'bars': '🍹 Bars',
    # Détente
    'parks': '🌳 Parcs',
    'beaches': '🏖️ Plages',
    'lakes': '🏞️ Lacs',
    # Divertissements
    'concert_halls': '🎶 Salles de concert',
    'nightclubs': '💃 Boîtes de nuit',
    'movie_theaters': '🎬 Cinémas',
    'stadiums': '🏟️ Stades',
    # Shopping
    'markets': '🌽 Marchés',
    'boutiques': '🛍️ Boutiques',
    'supermarkets': '🛒 Supermarchés',
    # Activités
    'festivals': '🎉 Festivals',
    'amusement_parks': "🎢 Parcs d'attractions",
    'zoos': '🐘 Zoos',
    'aquariums': '🐠 Aquariums',
    'mountain_resorts': '🏔️ Stations de montagne',
    # Sport
    'bike_rentals': '🚴 Locations de vélos',
    'campgrounds': '🏕️ Campings',
    'sports_centers': '🏋️‍♂️ Centres sportifs',
    'spas': '💆‍♀️ Spas',
    'gym': '🏋️‍♀️ Salles de sport',
    # Transports
    'train_stations': '🚆 Gares',
    'airports': '✈️ Aéroports',
    # Éducation
    'schools': '🏫 Écoles',
    # Santé
    'hospitals': '🏥 Hôpitaux',
}


def fetch_all_places(api_key, location):
    """Récupère tous les lieux par catégorie (1 appel searchText chacune) et
    applique le filtre pays. Séparé de l'écriture Excel pour que
    l'enrichissement Wikipédia (extraits de descriptions) passe AVANT."""
    location = " ".join(location.split())
    places_by_category = {}
    for category, description in CATEGORIES.items():
        print(f"Fetching data for category: {category}")
        data = search_places(api_key, location, category)
        if data:
            label = description.split(' ', 1)[1] if ' ' in description else description
            for p in data:
                p['_category'] = label
            places_by_category[category] = data
    return filter_by_country(places_by_category, location)


def create_excel_file(api_key, location, vacation_month, places_by_category=None):
    normalize_location(location)  # validation
    location = " ".join(location.split())
    vacation_month = " ".join(vacation_month.split()).replace(",", "-").replace(" ", "")
    sanitized_location = location.replace(",", "-").replace(" ", "")

    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_name = f"{sanitized_location}_{vacation_month}.xlsx"
    file_path = os.path.join(script_dir, file_name)

    # 1) + 2) Lieux déjà fournis (flux principal : fetch -> wiki -> Excel),
    # sinon récupération ici (compatibilité).
    if places_by_category is None:
        places_by_category = fetch_all_places(api_key, location)

    # 3) Écriture Excel (colonnes techniques retirées, classeur inchangé)
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    _write_legend(writer, 'catalog')
    for category, data in places_by_category.items():
        if not data:
            continue
        description = CATEGORIES[category]
        df = pd.DataFrame(data).sort_values(by='Total Reviews', ascending=False)
        df = df.drop(columns=['PlaceID', 'Lat', 'Lng', '_category',
                              'Country', 'CountryCode', 'Types', 'PrimaryType',
                              'BusinessStatus', 'WikiTitle', 'WikiLang',
                              'WikiViews', 'WikiUrl', 'WikiExtract'],
                     errors='ignore')
        sheet_name = description if len(description) <= 31 else description[:31]
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    writer.close()

    adjust_column_width(file_path)
    print(f"✅ Excel file created with clickable links: {file_path}")
    return file_path, location, vacation_month, places_by_category


def _write_legend(writer, kind):
    """Écrit une feuille 'Légende' en tête du classeur, qui explique les colonnes,
    les emojis et les liens. kind='catalog' (Excel principal) ou 'plan' (multi-jours)."""
    if kind == 'plan':
        rows = [
            ("Vue d'ensemble", "Récap des jours : nb d'étapes, distance, temps de marche, et le lien du parcours."),
            ("Une feuille par jour", "Les étapes dans l'ordre de marche optimisé (trajet le plus court)."),
            ("Étape 🎯", "Visite / lieu touristique."),
            ("Étape 🍴", "Déjeuner, inséré à mi-parcours."),
            ("Étape 🍽️", "Dîner, en fin de journée."),
            ("Étape 🍹", "Bar / sortie, après le dîner."),
            ("Total Reviews", "Nombre d'avis Google = popularité. Colonne en dégradé : plus c'est vert, plus c'est populaire."),
            ("Rating (on 5)", "Note moyenne Google sur 5."),
            ("Maps", "Lien cliquable vers la fiche Google Maps du lieu."),
            ("Wikipédia", "Article Wikipédia du lieu (quand reconnu automatiquement)."),
            ("Fermé", "Jour(s) de fermeture hebdomadaire - à vérifier avant d'y aller."),
            ("Parcours (Vue d'ensemble)", "Ouvre TOUT l'itinéraire du jour à pied dans Google Maps."),
        ]
    else:
        rows = [
            ("Feuilles", "Une feuille par catégorie (monuments, restaurants, plages, musées...)."),
            ("Total Reviews", "Nombre d'avis Google = popularité. Colonne en dégradé : plus c'est vert, plus c'est populaire."),
            ("Rating (on 5)", "Note moyenne Google sur 5."),
            ("Price Level", "Niveau de prix : Free, +, ++, +++, ++++."),
            ("Maps", "Lien cliquable vers la fiche Google Maps du lieu."),
            ("Website", "Lien cliquable vers le site officiel du lieu (quand renseigné)."),
            ("Fermeture", "Jour(s) de fermeture hebdomadaire du lieu (vide = ouvert 7j/7 ou horaires inconnus)."),
            ("Description", "Dernière colonne : résumé éditorial Google, enrichi par l'introduction Wikipédia quand l'article du lieu est reconnu ; à défaut, le type du lieu."),
            ("Astuce", "Chaque feuille est triée par popularité : les incontournables sont en haut."),
            ("Astuce", "Un plan multi-jours détaillé peut être généré (champ 'Nombre de jours' au lancement)."),
        ]
    df = pd.DataFrame(rows, columns=["Élément", "Signification"])
    df.to_excel(writer, sheet_name="Légende", index=False)


def create_multiday_excel(plan, location, vacation_month, script_dir):
    """Génère un 2e Excel : un plan sur plusieurs jours (1 feuille par jour +
    une feuille 'Vue d'ensemble'). Même style que le classeur principal.
    `plan` vient de build_multiday_plan. Renvoie le chemin du fichier ou None."""
    if not plan:
        return None
    sanitized = location.replace(",", "-").replace(" ", "")
    file_path = os.path.join(
        script_dir, f"{sanitized}_Plan-{len(plan)}jours_{vacation_month}.xlsx")
    writer = pd.ExcelWriter(file_path, engine="openpyxl")

    _write_legend(writer, 'plan')

    # Feuille récap : 1 ligne par jour + lien du parcours complet
    overview = []
    for d in plan:
        route = (f'=HYPERLINK("{d["url"]}", "🗺️ Ouvrir le parcours")'
                 if d.get("url") else "")
        overview.append({
            "Jour": d["day"],
            "Étapes": len(d["sequence"]),
            "Distance": d.get("distance_txt") or "-",
            "Marche": d.get("duration_txt") or "-",
            "Parcours": route,
        })
    pd.DataFrame(overview).to_excel(writer, sheet_name="Vue d'ensemble", index=False)

    # Une feuille par jour, dans l'ordre des étapes
    for d in plan:
        rows = []
        for i, s in enumerate(d["sequence"], start=1):
            wiki = (f'=HYPERLINK("{s["WikiUrl"]}", "📖 Wikipédia")'
                    if s.get("WikiUrl") else "")
            rows.append({
                "Ordre": i,
                "Étape": s.get("_step", ""),
                "Name": s.get("Name", ""),
                "Catégorie": s.get("_category", ""),
                "Rating (on 5)": s.get("Rating (on 5)", ""),
                "Total Reviews": s.get("Total Reviews", 0),
                "Maps": s.get("Maps", ""),
                "Wikipédia": wiki,
                "Fermé": s.get("Fermeture", ""),
                "Address": s.get("Address", ""),
                "Description": s.get("Description", ""),
            })
        pd.DataFrame(rows).to_excel(writer, sheet_name=f"Jour {d['day']}", index=False)
    writer.close()

    adjust_column_width(file_path)
    style_workbook(file_path)
    print(f"✅ Plan multi-jours créé : {file_path}")
    return file_path


# ----------------------------------------------------------------------------
# Email HTML (même contenu, même liens - juste stylé)
# ----------------------------------------------------------------------------
def build_email_bodies(location, location_cleaned, vacation_month_cleaned, itinerary_block="", itinerary_plain=""):
    ia_prompt = (
        f"Je prépare un voyage : {location}, en {vacation_month_cleaned.replace('-', ' ')}. "
        "J'ai DÉJÀ un guide Excel des meilleurs lieux classés par catégorie "
        "(monuments, plages, restaurants, musées...) : ne refais pas cette liste. "
        "Ta mission : la COMPLÉTER avec tout ce qu'un fichier statique ne peut pas "
        "donner, en faisant de VRAIES recherches web au moment où tu lis ceci, et "
        "livrer le résultat en ARTEFACT HTML. "
        "RECHERCHES (obligatoire) : croise au moins deux sources par information ; "
        "priorise l'office de tourisme et les sites officiels (mairie, billetteries, "
        "transporteurs, sites des lieux eux-mêmes) ; vérifie que chaque date tombe "
        "bien pendant mon séjour : rien de passé, rien de générique ni "
        "d'intemporel ; si une info (date, prix, horaire, dispo) n'est pas sûre, "
        "dis-le explicitement. "
        "CONTENU À COUVRIR : "
        "1) L'AGENDA du séjour : festivals, concerts, fêtes locales et de village, "
        "marchés (nocturnes inclus), expositions temporaires, événements sportifs : "
        "8 à 15 recommandations classées de la meilleure à la moins prioritaire, "
        "avec 2-3 coups de coeur mis en avant ; pour chacune : titre, date et "
        "horaire précis, lieu, prix approximatif, l'ambiance qui lui va (solo, "
        "couple, famille, amis), 2-3 phrases concrètes qui donnent envie, et un "
        "lien officiel ou de billetterie. "
        "2) EXCURSIONS : top 10 des villages, villes ou sites naturels autour, avec "
        "temps de trajet réaliste depuis ma base, le meilleur moyen d'y aller, et "
        "ce qui vaut vraiment le détour sur place. "
        "3) PRATIQUE : comment y aller et se déplacer sur place (options, pièges du "
        "type parking ou navette ou réservation obligatoire), budget réaliste par "
        "poste (repas, activités, transports) pour la période, météo normale du "
        "mois et quoi mettre dans la valise, affluence attendue et conseils de "
        "réservation, documents nécessaires, numéros d'urgence, arnaques courantes, "
        "santé. "
        "4) GASTRONOMIE et COUTUMES : spécialités locales à goûter avec 3 à 5 "
        "adresses vérifiées, usages locaux et astuces pour profiter comme un "
        "habitant. "
        "ARTEFACT HTML (contraintes techniques obligatoires) : un SEUL fichier HTML "
        "autonome, CSS en ligne, AUCUN script, AUCUNE ressource externe (ni iframe, "
        "ni police, ni image distante, ni localStorage). Échappe TOUTES les "
        "esperluettes en entité HTML, dans les URL comme dans les textes. Liens "
        "complets en https, guillemets droits pour les attributs, target blank. "
        "Structure : en-tête résumant le séjour, section Agenda en cartes (badge "
        "pour les coups de coeur, date et horaire, lieu, prix, bouton vers le lien "
        "officiel, bouton Itinéraire au format officiel Google Maps api=1 en "
        "échappant chaque esperluette), section Excursions avec temps de trajet, "
        "puis les sections pratiques compactes en tableaux. "
        "TERMINE l'artefact par une rangée de 4 boutons cliquables pour rebondir : "
        "'Agenda semaine par semaine', 'Excursions et villages', 'Budget et bons "
        "plans', 'Gastronomie locale' : chaque bouton est un lien https vers "
        "claude.ai/new avec le paramètre q contenant une nouvelle recherche ciblée "
        "qui rappelle la destination et la période. "
        "VÉRIFICATION DES LIENS : chaque lien pointe vers une page officielle qui "
        "existe réellement, vérifiée par une recherche web ; ne devine jamais une "
        "URL. "
        "DOUBLE VÉRIFICATION AVANT DE LIVRER (indique le résultat en une seule "
        "ligne) : 1) balises html, head, style et body équilibrées ; 2) zéro "
        "esperluette non échappée ; 3) tous les liens complets et vérifiés ; "
        "4) aucun script ni ressource externe. Si un point échoue, corrige puis "
        "re-vérifie avant de livrer. "
        "SÉCURITÉ D'AFFICHAGE : fournis aussi le fichier téléchargeable en plus de "
        "l'aperçu, et rappelle en une ligne : si l'aperçu reste blanc, basculer "
        "entre la vue code et la vue aperçu, ou ouvrir le fichier dans le "
        "navigateur. Reste concret, va droit au but."
    )

    import urllib.parse as _up
    claude_url_plain = "https://claude.ai/new?q=" + _up.quote(ia_prompt)

    # --- Version texte brut (fallback, identique à l'originale) ---
    body_plain = f"""Bienvenue à bord de Place Explorer !

Découvrez {location} ! C'est une destination rêvée pour des aventures inoubliables. Votre guide inclut :
    - Les lieux incontournables à visiter
    - Un plan d'organisation pour votre voyage

{itinerary_plain}🗺 Étapes pour organiser votre voyage :
    1. Ouvrez le fichier Excel attaché avec Google Sheet en cliquant une fois dessus
    2. Consultez chaque feuille pour explorer les meilleurs options par catégorie
    3. Planifiez vos activités (par exemple 2/jours par feuille) sur : https://www.google.com/mymaps un calque par ville
    4. Si le réseau est payant à l’étranger, utilisez Google Maps hors connexion : https://support.google.com/maps/answer/6291838?hl=fr
    5. Envoyez-vous votre parcours du jour (en fonction de la proximité) sur WhatsApp ou via Google Docs pour le garder à portée de main sur votre téléphone : https://web.whatsapp.com ou https://docs.google.com

🌍 Liens utiles :
    - ✈ Pour trouver les vols les moins chers et obtenir des indemnisations en cas de retard : https://www.skyscanner.fr/ ou https://www.airhelp.com/fr/
    - 🚅 Pour comparer tous les moyens de transport : https://www.rome2rio.com/
    - 🏠 Pour réserver votre hébergement et véhicule : https://www.airbnb.com et https://www.booking.com
    - 🖍 Pour des avis et recommandations : https://www.tripadvisor.com
    - 🗺 Pour réserver des activités locales : https://www.getyourguide.fr/
    - 📞 Pour trouver des eSIM à moindre coût à l'étranger : https://www.airalo.com/fr
    - 💳 Pour dépenser sans frais de change à l'étranger, ouvrez un compte Revolut : https://revolut.com/referral/?referral-code=romainavh3!DEC1-24-VR-FR

🤖 Ouvrez ce lien pour lancer le prompt ci-dessous dans Claude (déjà pré-rempli, appuyez juste sur Entrée) : {claude_url_plain}\n\nOu copiez-le manuellement sur https://claude.ai :

"{ia_prompt}"

Nous espérons que vous passerez un moment incroyable. Bon voyage ! ✈
N'hésitez pas à faire un don via PayPal à l'adresse romtaug@gmail.com si cela vous a aidé.
Accédez à notre outil pour travailler à l'étranger : https://bordeuroconnect.netlify.app/"""

    # --- Version HTML professionnelle (mêmes liens, même contenu) ---
    NAVY = '#1f3b63'
    RED = '#d32f2f'
    TEXT = '#3c4043'
    MUTED = '#70757a'

    def btn(url, label, bg=NAVY):
        return (f'<a href="{url}" target="_blank" '
                f'style="display:inline-block;background-color:{bg};color:#ffffff;'
                f'padding:11px 22px;border-radius:6px;text-decoration:none;font-weight:600;'
                f'font-size:13px;letter-spacing:0.2px;">{label}</a>')

    def section_title(label, title):
        return (f'<p style="margin:0 0 2px;font-size:11px;font-weight:700;letter-spacing:1.5px;'
                f'text-transform:uppercase;color:{RED};">{label}</p>'
                f'<h2 style="margin:0 0 16px;font-size:19px;font-weight:700;color:{NAVY};">{title}</h2>')

    def step(num, text, buttons=''):
        badge = (f'<span style="display:inline-block;width:26px;height:26px;line-height:26px;'
                 f'border-radius:50%;background-color:{NAVY};color:#ffffff;text-align:center;'
                 f'font-size:13px;font-weight:700;">{num}</span>')
        extra = f'<div style="margin-top:8px;">{buttons}</div>' if buttons else ''
        return (f'<tr><td style="padding:10px 14px 10px 0;vertical-align:top;width:26px;">{badge}</td>'
                f'<td style="padding:12px 0;font-size:14px;color:{TEXT};line-height:1.6;">{text}{extra}</td></tr>')

    def useful(name, text, buttons):
        return (f'<tr><td style="padding:14px 0;border-bottom:1px solid #eceff1;">'
                f'<p style="margin:0 0 2px;font-size:14px;font-weight:700;color:{NAVY};">{name}</p>'
                f'<p style="margin:0 0 9px;font-size:13px;color:{TEXT};line-height:1.55;">{text}</p>'
                f'{buttons}</td></tr>')

    def small_btn(url, label):
        return (f'<a href="{url}" target="_blank" '
                f'style="display:inline-block;background-color:#ffffff;color:{NAVY};'
                f'border:1.5px solid {NAVY};padding:8px 16px;border-radius:6px;text-decoration:none;'
                f'font-weight:600;font-size:12.5px;margin:0 8px 6px 0;">{label} &rarr;</a>')

    preheader = (f"Votre guide de voyage personnalisé pour {location} : lieux incontournables, "
                 f"plan d'organisation et liens utiles.")

    # URL Claude avec le prompt pré-rempli dans la zone de saisie (le destinataire n'a plus qu'à appuyer sur Entrée)
    import urllib.parse
    claude_url = "https://claude.ai/new?q=" + urllib.parse.quote(ia_prompt)

    body_html = f"""\
<!DOCTYPE html>
<html lang="fr">
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background-color:#eef1f4;font-family:'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
<div style="display:none;max-height:0;overflow:hidden;mso-hide:all;">{preheader}</div>
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background-color:#eef1f4;padding:32px 12px;">
<tr><td align="center">
<table role="presentation" width="620" cellpadding="0" cellspacing="0" style="max-width:620px;width:100%;background:#ffffff;border-radius:14px;overflow:hidden;box-shadow:0 2px 10px rgba(31,59,99,0.10);">

  <!-- En-tête de marque -->
  <tr><td style="padding:28px 36px 20px;text-align:center;">
    <img src="cid:logo" alt="Place Explorer" width="72" style="display:block;margin:0 auto;max-width:72px;height:auto;">
    <p style="margin:10px 0 0;font-size:13px;font-weight:700;letter-spacing:3px;text-transform:uppercase;color:{NAVY};">Place&nbsp;Explorer</p>
  </td></tr>

  <!-- Bannière -->
  <tr><td style="padding:0;">
    <img src="cid:travel" alt="Préparation de voyage" width="620" style="display:block;width:100%;height:auto;">
  </td></tr>

  <!-- Titre -->
  <tr><td style="padding:32px 36px 6px;text-align:center;">
    <span style="display:inline-block;background-color:#fdecea;color:{RED};font-size:11.5px;font-weight:700;letter-spacing:1px;text-transform:uppercase;padding:6px 14px;border-radius:20px;margin-bottom:12px;">Guide voyage &middot; {vacation_month_cleaned.replace('-', ' ')}</span>
    <h1 style="margin:0 0 8px;font-size:25px;font-weight:700;color:{NAVY};line-height:1.3;">Votre guide de voyage pour {location}</h1>
    <p style="margin:0;font-size:15px;color:{TEXT};line-height:1.6;">
      Bienvenue à bord&nbsp;! Découvrez <strong>{location}</strong>, une destination rêvée pour des aventures inoubliables.
      Votre guide inclut les lieux incontournables à visiter et un plan d'organisation pour votre voyage.
    </p>
  </td></tr>

  <!-- Encart pièce jointe -->
  <tr><td style="padding:20px 36px 8px;">
    <table role="presentation" cellpadding="0" cellspacing="0" width="100%" style="background-color:#f3f6fa;border:1px solid #dde5ee;border-radius:10px;">
      <tr>
        <td style="padding:16px 18px;font-size:24px;width:36px;vertical-align:middle;">📎</td>
        <td style="padding:16px 18px 16px 0;font-size:13.5px;color:{TEXT};line-height:1.55;vertical-align:middle;">
          <strong style="color:{NAVY};">Votre guide Excel est en pièce jointe.</strong><br>
          Plus de 30 catégories (monuments, restaurants, plages, musées…), top 20 des lieux par popularité, liens Google&nbsp;Maps cliquables.
        </td>
      </tr>
    </table>
  </td></tr>

  {itinerary_block}

  <!-- Étapes -->
  <tr><td style="padding:26px 36px 8px;">
    {section_title('Organisation', 'Étapes pour organiser votre voyage')}
    <table role="presentation" cellpadding="0" cellspacing="0" width="100%">
      {step('1', "Ouvrez le fichier Excel attaché avec <strong>Google&nbsp;Sheets</strong> en cliquant une fois dessus.")}
      {step('2', "Consultez chaque feuille pour explorer les meilleures options par catégorie.")}
      {step('3', "Planifiez vos activités (par exemple 2 par jour et par feuille), avec un calque par ville&nbsp;:",
            small_btn('https://www.google.com/mymaps', 'Google My Maps'))}
      {step('4', "Si le réseau est payant à l'étranger, utilisez Google Maps hors connexion&nbsp;:",
            small_btn('https://support.google.com/maps/answer/6291838?hl=fr', 'Maps hors connexion'))}
      {step('5', "Envoyez-vous votre parcours du jour (en fonction de la proximité) pour le garder à portée de main sur votre téléphone&nbsp;:",
            small_btn('https://web.whatsapp.com', 'WhatsApp Web') +
            small_btn('https://docs.google.com', 'Google Docs'))}
    </table>
  </td></tr>

  <!-- Liens utiles -->
  <tr><td style="padding:26px 36px 8px;">
    {section_title('Ressources', 'Liens utiles pour votre voyage')}
    <table role="presentation" cellpadding="0" cellspacing="0" width="100%">
      {useful('Vols', "Trouvez les vols les moins chers et obtenez des indemnisations en cas de retard.",
              small_btn('https://www.skyscanner.fr/', 'Skyscanner') + small_btn('https://www.airhelp.com/fr/', 'AirHelp'))}
      {useful('Transports', "Comparez tous les moyens de transport pour chaque trajet.",
              small_btn('https://www.rome2rio.com/', 'Rome2Rio'))}
      {useful('Hébergement &amp; véhicule', "Réservez votre logement et votre véhicule en quelques clics.",
              small_btn('https://www.airbnb.com', 'Airbnb') + small_btn('https://www.booking.com', 'Booking.com'))}
      {useful('Avis &amp; recommandations', "Consultez les avis des voyageurs avant de choisir.",
              small_btn('https://www.tripadvisor.com', 'Tripadvisor'))}
      {useful('Activités locales', "Réservez des visites et expériences sur place.",
              small_btn('https://www.getyourguide.fr/', 'GetYourGuide'))}
      {useful("Connexion à l'étranger", "Trouvez des eSIM à moindre coût pour rester connecté.",
              small_btn('https://www.airalo.com/fr', 'Airalo'))}
    </table>
  </td></tr>

  <!-- Offre Revolut -->
  <tr><td style="padding:24px 36px 8px;">
    <table role="presentation" cellpadding="0" cellspacing="0" width="100%" style="background-color:{NAVY};border-radius:12px;">
      <tr><td style="padding:24px 26px;text-align:center;">
        <p style="margin:0 0 4px;font-size:11px;font-weight:700;letter-spacing:1.5px;text-transform:uppercase;color:#9fb3d1;">Offre partenaire</p>
        <p style="margin:0 0 6px;font-size:17px;font-weight:700;color:#ffffff;">Dépensez sans frais de change à l'étranger</p>
        <p style="margin:0 0 14px;font-size:13.5px;color:#d7e0ec;line-height:1.55;">Ouvrez un compte Revolut en quelques minutes.</p>
        {btn('https://revolut.com/referral/?referral-code=romainavh3!DEC1-24-VR-FR', 'Ouvrir un compte Revolut', RED)}
      </td></tr>
    </table>
  </td></tr>

  <!-- Prompt Claude -->
  <tr><td style="padding:26px 36px 8px;">
    {section_title('Assistant IA', 'Enrichissez votre expérience')}
    <p style="margin:0 0 12px;font-size:14px;color:{TEXT};line-height:1.6;">
      Cliquez sur le bouton ci-dessous&nbsp;: <strong>Claude</strong> s'ouvre avec ce prompt déjà saisi, il ne reste qu'à appuyer sur Entrée. Il fait alors de vraies recherches web du moment et livre le complément vivant de votre Excel&nbsp;: agenda daté du séjour, excursions, budget, météo et conseils, en page HTML interactive&nbsp;:
    </p>
    <table role="presentation" cellpadding="0" cellspacing="0" width="100%" style="background-color:#f8f9fa;border-left:4px solid {NAVY};border-radius:0 8px 8px 0;">
      <tr><td style="padding:16px 18px;">
        <p style="margin:0;font-size:12.5px;color:#444746;line-height:1.65;font-family:Consolas,'Courier New',monospace;">"{ia_prompt}"</p>
      </td></tr>
    </table>
    <div style="text-align:center;margin-top:16px;">
      {btn(claude_url, 'Ouvrir Claude - prompt pré-rempli')}
    </div>
  </td></tr>

  <!-- Bon voyage -->
  <tr><td style="padding:28px 36px 24px;text-align:center;">
    <p style="margin:0;font-size:15px;color:{NAVY};font-weight:700;">Nous espérons que vous passerez un moment incroyable. Bon voyage&nbsp;!</p>
  </td></tr>

  <!-- Pied de page -->
  <tr><td style="padding:26px 36px 30px;background-color:#f7f9fb;border-top:1px solid #e8ecf1;text-align:center;">
    <p style="margin:0 0 10px;font-size:13px;color:{TEXT};line-height:1.6;">
      Ce guide vous a été utile&nbsp;? Soutenez le projet d'un don via <strong>PayPal</strong>
      à l'adresse <a href="mailto:romtaug@gmail.com" style="color:{NAVY};font-weight:600;text-decoration:none;">romtaug@gmail.com</a>
      ou en scannant ce QR&nbsp;code&nbsp;:
    </p>
    <img src="cid:qrcode" alt="QR code don PayPal" width="110" style="display:block;margin:8px auto 14px;max-width:110px;height:auto;border-radius:8px;border:1px solid #e0e5ea;">
    <p style="margin:0 0 10px;font-size:13px;color:{TEXT};">Découvrez aussi notre outil pour travailler à l'étranger&nbsp;:</p>
    {small_btn('https://bordeuroconnect.netlify.app/', 'BordEuro Connect')}
    <p style="margin:18px 0 0;font-size:11.5px;color:{MUTED};line-height:1.6;">
      Vous recevez cet e-mail car un guide Place&nbsp;Explorer a été généré pour vous.<br>
      Place&nbsp;Explorer - Guide de voyage automatisé
    </p>
  </td></tr>

</table>
</td></tr>
</table>
</body>
</html>"""

    return body_plain, body_html


def send_email_with_excel(sender_email, password, receiver_emails, subject,
                          body_plain, body_html, file_path, image_paths,
                          extra_files=None):
    msg = MIMEMultipart('mixed')
    msg['From'] = sender_email
    msg['To'] = ", ".join(receiver_emails)
    msg['Subject'] = subject

    # Corps : alternative (texte brut + HTML avec images inline)
    alt = MIMEMultipart('alternative')
    alt.attach(MIMEText(body_plain, 'plain', 'utf-8'))

    related = MIMEMultipart('related')
    related.attach(MIMEText(body_html, 'html', 'utf-8'))

    # Images intégrées au corps HTML (cid:logo, cid:travel, cid:qrcode)
    for image_path in image_paths:
        cid = os.path.splitext(os.path.basename(image_path))[0]  # logo / travel / qrcode
        if os.path.exists(image_path):
            with open(image_path, 'rb') as img:
                img_part = MIMEImage(img.read())
            img_part.add_header('Content-ID', f'<{cid}>')
            img_part.add_header('Content-Disposition', 'inline',
                                filename=os.path.basename(image_path))
            related.attach(img_part)
        else:
            print(f"Image non trouvée (ignorée) : {image_path}")

    alt.attach(related)
    msg.attach(alt)

    # Pièces jointes Excel (classeur principal + éventuel plan multi-jours)
    for path in [file_path] + list(extra_files or []):
        if not path:
            continue
        try:
            with open(path, 'rb') as file:
                part = MIMEBase('application', "octet-stream")
                part.set_payload(file.read())
            encoders.encode_base64(part)
            normalized_name = normalize_filename(os.path.basename(path))
            part.add_header('Content-Disposition', f'attachment; filename="{normalized_name}"')
            msg.attach(part)
        except FileNotFoundError:
            print(f"Fichier non trouvé : {path}")
            if path == file_path:
                return

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender_email, password)
            server.sendmail(sender_email, receiver_emails, msg.as_string())
            print(f"✅ E-mail envoyé avec succès à {', '.join(receiver_emails)}")
    except Exception as e:
        print(f"❌ Erreur lors de l'envoi de l'e-mail : {e}")
        sys.exit(1)


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
if __name__ == "__main__":
    try:
        # 1) Récupération des lieux (Places New, 1 appel/catégorie)
        places_by_category = fetch_all_places(API_KEY, LOCATION)
        # 2) Wikipédia AVANT l'Excel : articles + extraits d'introduction,
        #    fusionnés dans la colonne Description. WIKI_ENRICH=0 désactive.
        if os.environ.get("WIKI_ENRICH", "1") != "0":
            try:
                enrich_with_wikipedia(places_by_category)
                finalize_descriptions(places_by_category)
            except Exception as e:
                print(f"⚠️ Enrichissement Wikipédia ignoré : {e}")
        # 3) Écriture + style de l'Excel
        excel_file, location, vacation_month, places_by_category = create_excel_file(
            API_KEY, LOCATION, VACATION_MONTH, places_by_category=places_by_category)
        style_workbook(excel_file)
    except SystemExit:
        raise
    except Exception as e:
        print(f"❌ Erreur lors de la création du fichier Excel : {e}")
        sys.exit(1)

    location_cleaned = location.replace(",", "-").replace(" ", "")
    vacation_month_cleaned = vacation_month.replace(" ", "-")

    if not os.path.exists(excel_file):
        print(f"❌ Le fichier Excel n'a pas été trouvé : {excel_file}")
        sys.exit(1)

    # Garde-fou : vérifier que l'Excel contient des données avant d'envoyer
    wb_check = load_workbook(excel_file, read_only=True)
    total_rows = 0
    for sheet_name in wb_check.sheetnames:
        if sheet_name == "Légende":      # feuille explicative -> ne compte pas comme données
            continue
        n = sum(1 for _ in wb_check[sheet_name].iter_rows(min_row=2))
        total_rows += n
        print(f"   📄 {sheet_name} : {n} lieux")
    wb_check.close()
    print(f"📊 Total : {total_rows} lieux dans {len(wb_check.sheetnames)} feuilles")
    if total_rows == 0:
        print("❌ Le fichier Excel est vide (0 lieu) - envoi annulé.")
        print("   Causes probables : statut API en erreur (voir ⚠️ ci-dessus) ou filtre pays trop strict.")
        sys.exit(1)

    script_dir = os.path.dirname(os.path.abspath(__file__))
    image_paths = [
        os.path.join(script_dir, "Image", "logo.png"),
        os.path.join(script_dir, "Image", "travel.png"),
        os.path.join(script_dir, "Image", "qrcode.png"),
    ]

    # --- Itinéraire du jour : UNIQUEMENT pour une ville ---------------------
    # Une virgule dans la destination = format "Ville, Pays" -> on génère le
    # parcours à pied. Sans virgule = pays entier -> pas d'itinéraire : il
    # s'ancrerait sur un point arbitraire (ex. Trakai au lieu de Vilnius) et
    # perdrait souvent le déjeuner / dîner / bar. L'Excel, lui, reste complet.
    itinerary_block, itinerary_plain = "", ""
    multiday_file = None
    is_city = "," in location
    if is_city:
        # Ancre fiable = coordonnées de la ville demandée (et non le lieu le plus
        # visité, qui peut être dans une autre ville). Si le géocodage échoue,
        # build_day_itinerary retombe sur la zone la plus dense des résultats.
        city_anchor = geocode_location(location, API_KEY)
        if city_anchor:
            print(f"📍 Ancre itinéraire (ville géocodée) : {city_anchor[0]:.4f}, {city_anchor[1]:.4f}")
        # Points de vue OSM injectés APRÈS l'Excel (parcours uniquement) ; ils
        # récupèrent leurs vues/extraits Wikipédia à la volée via wiki_enrich.
        if city_anchor and os.environ.get("OSM_VIEWPOINTS", "1") != "0":
            inject_osm_viewpoints(places_by_category, city_anchor)
            if os.environ.get("WIKI_ENRICH", "1") != "0":
                for _vp in places_by_category.get('viewpoints', []):
                    try:
                        wiki_enrich(_vp)
                    except Exception:
                        pass
        itin = build_day_itinerary(places_by_category, api_key=API_KEY, anchor=city_anchor)
        if itin:
            print("🗺️ Itinéraire du jour :")
            print(describe(itin["sequence"]))
            if itin.get("distance_txt"):
                print(f"   {itin['distance_txt']} · {itin['duration_txt']} de marche")
            print(itin["url"])
            map_path = os.path.join(script_dir, "routemap.png")
            if download_static_map(itin, API_KEY, map_path):
                itin["has_map"] = True
                image_paths.append(map_path)   # cid:routemap -> aperçu dans l'email
            itinerary_block = itinerary_email_block(itin)
            itinerary_plain = itinerary_plain_block(itin)

        # Plan multi-jours (Excel séparé) si un nombre de jours est demandé
        if NB_DAYS and NB_DAYS >= 1:
            plan = build_multiday_plan(places_by_category, NB_DAYS,
                                       anchor=city_anchor, api_key=API_KEY)
            multiday_file = create_multiday_excel(plan, location, vacation_month, script_dir)
    else:
        print("ℹ️ Destination = pays entier -> pas d'itinéraire ni de plan multi-jours "
              "(les parcours à pied n'ont de sens que pour une ville). "
              "Précise une ville (ex : 'Vilnius, Lithuania') pour les obtenir.")

    subject = f"🌍 PlaceExplorer : Les Meilleurs Lieux Destination {location_cleaned} en {vacation_month_cleaned}"
    body_plain, body_html = build_email_bodies(location, location_cleaned, vacation_month_cleaned,
                                               itinerary_block=itinerary_block, itinerary_plain=itinerary_plain)

    send_email_with_excel(SENDER_EMAIL, GMAIL_APP_PASSWORD, RECEIVER_EMAILS,
                          subject, body_plain, body_html, excel_file, image_paths,
                          extra_files=[multiday_file] if multiday_file else None)

    # Déplacer l'Excel dans Content/ (récupéré ensuite comme artifact par le workflow)
    content_dir = os.path.join(script_dir, "Content")
    os.makedirs(content_dir, exist_ok=True)
    destination = os.path.join(content_dir, os.path.basename(excel_file))
    shutil.move(excel_file, destination)
    print(f"✅ Fichier déplacé dans le dossier Content : {destination}")

    if multiday_file and os.path.exists(multiday_file):
        dest2 = os.path.join(content_dir, os.path.basename(multiday_file))
        shutil.move(multiday_file, dest2)
        print(f"✅ Plan multi-jours déplacé dans Content : {dest2}")
